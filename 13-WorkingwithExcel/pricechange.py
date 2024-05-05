import openpyxl 
from openpyxl.styles import Font

wb = openpyxl.load_workbook("/Users/farihaalam/produceSaless.xlsx")
sheet = wb['Sheet']

Price_updates = {'Garlic':3.07,
                 'Celery':1.19,
                 'Lemon':1.27}

for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in Price_updates:
        sheet.cell(row=rowNum, column=2).value = Price_updates[produceName]

wb.save("/Users/farihaalam/updatedProducesales.xlsx")