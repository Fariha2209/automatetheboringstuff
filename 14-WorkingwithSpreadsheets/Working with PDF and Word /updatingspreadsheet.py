import openpyxl
import openpyxl.workbook 

wb = openpyxl.load_workbook('/Users/farihaalam/produceSalescopy.xlsx')
sheet = wb['Sheet']

price_mapping = {'Celery': 1.19, 
                 'Garlic': 3.07, 
                 'Lemon': 1.27}

#print(column_A_values)
# print(column_B_values)
# print(column_C_values)
# print(column_D_values)

for rowNum in range(2, sheet.max_row):
  produce_name = sheet.cell(row=rowNum, column=1).value #check produce name in columnA
  if produce_name in price_mapping:
    sheet.cell(row=rowNum, column=2).value = price_mapping[produce_name]

   #print('hello')
     
wb.save('/Users/farihaalam/produceSalescopy5.xlsx')




 
